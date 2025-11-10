"""document_processor.py
Core ingestion & indexing pipeline for the RAG system.

High-level responsibilities:
        1. Load raw source files of multiple types (PDF, TXT, Word, Excel).
        2. Normalize them into `Document` objects with consistent metadata.
        3. Split ("chunk") large texts into smaller semantic units for precise retrieval.
        4. Generate embeddings for each chunk using the configured embedding model.
        5. Persist embeddings & metadata in a vector store (Chroma) for future similarity search.

Why chunking & embeddings:
        • Large documents exceed typical LLM context limits; chunking lets us select only
            the most relevant pieces.
        • Embeddings map text → high-dimensional vectors where semantic similarity becomes
            numeric distance; this enables efficient top‑K retrieval.

Design decisions:
        • RecursiveCharacterTextSplitter with configurable `CHUNK_SIZE` & `CHUNK_OVERLAP`
            balances precision vs. continuity (overlap preserves sentence/paragraph flow).
        • Persistent Chroma store avoids recomputation of embeddings between runs.
        • File‑type specific loaders (PDF/Word/Excel) maintain flexibility and richer metadata.
        • Optional Excel handling flattens sheets into pipe‑separated rows so tabular content
            becomes searchable text.

Public API surface:
        - `process_all_documents()`: One-shot ingestion + chunk + index routine.
        - `search(query, k)`: Similarity search returning top‑K matching chunks.
        - `get_stats()`: Lightweight health/introspection of the current vector store.
        - `add_documents_to_vectorstore(chunks)`: Internal helper exposed for incremental updates.

Error & safety notes:
        • Vector store initialization may fail (disk, schema, permissions); guarded & logged.
        • Adding documents checks that the vector store is not None before proceeding.
        • Unsupported file types are skipped with a warning instead of raising.

Extensibility hooks:
        • Add new loaders (e.g., Markdown) by extending `load_document`.
        • Swap vector store implementation by adjusting `_load_or_create_vectorstore` and methods.
        • Apply similarity score filtering or metadata filters in `search` if needed for precision.

Typical usage:
        >>> processor = DocumentProcessor()
        >>> stats = processor.process_all_documents()
        >>> results = processor.search("What is retrieval augmented generation?", k=4)
        >>> for doc in results: print(doc.page_content[:200])

All configuration constants (chunk size, embedding model, paths) live in `config.py`.
Modify them there; re-run `process_all_documents()` to re-index after changing chunking.
"""

import logging
import os
from pathlib import Path
from typing import Dict, List, Optional

# LangChain imports
from langchain_community.document_loaders import (
    PyPDFLoader, TextLoader, UnstructuredWordDocumentLoader)
from langchain_community.embeddings import OllamaEmbeddings
from langchain_community.vectorstores import Chroma
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

import fix_sqlite
# Our config
from config import (CHUNK_OVERLAP, CHUNK_SIZE, DOCUMENTS_DIR, EMBEDDING_MODEL,
                    OLLAMA_BASE_URL, VECTORSTORE_DIR)

# Set up logging so we can see what's happening
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DocumentProcessor:
    """
    Handles all document processing for our RAG system.

    Think of this as your librarian who:
    1. Reads books (loads documents)
    2. Makes index cards (creates chunks)
    3. Organizes them (creates embeddings)
    4. Files them away (stores in vector DB)
    """

    def __init__(self):
        """Initialize the document processor"""
        logger.info("Initializing Document Processor...")

        # Create embeddings model - this converts text to numbers
        self.embeddings = OllamaEmbeddings(
            model=EMBEDDING_MODEL, base_url=OLLAMA_BASE_URL
        )

        # Text splitter - breaks documents into chunks
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=CHUNK_SIZE,
            chunk_overlap=CHUNK_OVERLAP,
            length_function=len,
            separators=["\n\n", "\n", " ", ""],  # Try to split on paragraphs first
        )

        # Vector store - our searchable database (assigned in _load_or_create_vectorstore)
        # Annotated Optional so type checkers require a guard before use.
        self.vectorstore: Optional[Chroma] = None
        self._load_or_create_vectorstore()

        logger.info("✓ Document Processor ready!")

    def _load_or_create_vectorstore(self):
        """
        Load existing vector store or create a new one.
        Vector store persists to disk, so we don't re-process documents every time.
        """
        try:
            # Try to load existing vectorstore
            if (VECTORSTORE_DIR / "chroma.sqlite3").exists():
                logger.info("Loading existing vector store...")
                self.vectorstore = Chroma(
                    persist_directory=str(VECTORSTORE_DIR),
                    embedding_function=self.embeddings,
                )
                # Get count of documents
                count = self.vectorstore._collection.count()
                logger.info(f"✓ Loaded vector store with {count} chunks")
            else:
                logger.info("Creating new vector store...")
                self.vectorstore = Chroma(
                    persist_directory=str(VECTORSTORE_DIR),
                    embedding_function=self.embeddings,
                )
                logger.info("✓ Created new vector store")
        except Exception as e:
            logger.error(f"Error with vector store: {e}")
            raise

    def load_document(self, file_path: Path) -> List[Document]:
        """
        Load a single document based on its file type.

        Args:
            file_path: Path to the document

        Returns:
            List of LangChain Document objects

        The Document object has:
        - page_content: The actual text
        - metadata: Info about the document (filename, page number, etc.)
        """
        logger.info(f"Loading: {file_path.name}")

        file_extension = file_path.suffix.lower()

        try:
            if file_extension == ".pdf":
                loader = PyPDFLoader(str(file_path))
                documents = loader.load()

            elif file_extension == ".txt":
                loader = TextLoader(str(file_path), encoding="utf-8")
                documents = loader.load()

            elif file_extension in [".docx", ".doc"]:
                loader = UnstructuredWordDocumentLoader(str(file_path))
                documents = loader.load()

            elif file_extension in [".xlsx", ".xls"]:
                # For Excel, we'll read it differently
                documents = self._load_excel(file_path)

            else:
                logger.warning(f"Unsupported file type: {file_extension}")
                return []

            # Add filename to metadata
            for doc in documents:
                doc.metadata["source"] = file_path.name

            logger.info(f"✓ Loaded {len(documents)} page(s) from {file_path.name}")
            return documents

        except Exception as e:
            logger.error(f"Error loading {file_path.name}: {e}")
            return []

    def _load_excel(self, file_path: Path) -> List[Document]:
        """
        Special handler for Excel files.
        We'll convert sheets to text format.
        """
        import openpyxl

        documents = []
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Convert sheet to text
            text_content = f"Sheet: {sheet_name}\n\n"

            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if any(cell is not None for cell in row):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    text_content += row_text + "\n"

            # Create document
            doc = Document(
                page_content=text_content,
                metadata={"source": file_path.name, "sheet": sheet_name},
            )
            documents.append(doc)

        return documents

    def chunk_documents(self, documents: List[Document]) -> List[Document]:
        """
        Split documents into smaller chunks.

        Why chunk?
        - LLMs have limited context windows
        - Smaller chunks = more precise retrieval
        - Better to retrieve 4 relevant paragraphs than 4 entire documents

        Args:
            documents: List of loaded documents

        Returns:
            List of chunked documents
        """
        logger.info(f"Chunking {len(documents)} document(s)...")

        chunks = self.text_splitter.split_documents(documents)

        logger.info(f"✓ Created {len(chunks)} chunks")
        return chunks

    def add_documents_to_vectorstore(self, chunks: List[Document]):
        """
        Create embeddings and add to vector store.

        This is where the "magic" happens:
        1. Each chunk's text is converted to a vector (list of numbers)
        2. Similar text gets similar vectors
        3. These vectors are stored in ChromaDB
        4. Later, we can find similar vectors quickly!

        Args:
            chunks: List of document chunks
        """
        logger.info(f"Creating embeddings for {len(chunks)} chunks...")
        logger.info("(This may take a while...)")

        # Defensive check + narrowing for static analyzers (Pylance/mypy)
        if self.vectorstore is None:
            logger.error("Vector store not initialized. Aborting add.")
            raise RuntimeError("Vector store is not initialized")

        try:
            # This calls Ollama to create embeddings for each chunk
            self.vectorstore.add_documents(chunks)

            # Persist to disk
            self.vectorstore.persist()

            logger.info("✓ Documents added to vector store")

        except Exception as e:
            logger.error(f"Error adding documents: {e}")
            raise

    def process_all_documents(self) -> Dict[str, int]:
        """
        Process all documents in the documents folder.
        This is the main function you'll call to index your documents.

        Returns:
            Statistics about processed documents
        """
        logger.info("=" * 50)
        logger.info("PROCESSING ALL DOCUMENTS")
        logger.info("=" * 50)

        all_chunks = []
        stats = {"files_processed": 0, "files_failed": 0, "total_chunks": 0}

        # Get all files from documents directory
        supported_extensions = [".pdf", ".txt", ".docx", ".doc", ".xlsx", ".xls"]
        files = [
            f
            for f in DOCUMENTS_DIR.iterdir()
            if f.is_file() and f.suffix.lower() in supported_extensions
        ]

        if not files:
            logger.warning(f"No documents found in {DOCUMENTS_DIR}")
            return stats

        logger.info(f"Found {len(files)} document(s) to process")

        # Process each file
        for file_path in files:
            # Load document
            documents = self.load_document(file_path)

            if documents:
                # Chunk the document
                chunks = self.chunk_documents(documents)
                all_chunks.extend(chunks)
                stats["files_processed"] += 1
            else:
                stats["files_failed"] += 1

        # Add all chunks to vector store
        if all_chunks:
            self.add_documents_to_vectorstore(all_chunks)
            stats["total_chunks"] = len(all_chunks)

        logger.info("=" * 50)
        logger.info("PROCESSING COMPLETE")
        logger.info(f"✓ Files processed: {stats['files_processed']}")
        logger.info(f"✗ Files failed: {stats['files_failed']}")
        logger.info(f"📦 Total chunks created: {stats['total_chunks']}")
        logger.info("=" * 50)

        return stats

    def search(self, query: str, k: int = 4) -> List[Document]:
        """
        Search for relevant documents.

        This is the "Retrieval" part of RAG!

        Args:
            query: The user's question
            k: Number of results to return

        Returns:
            List of most relevant document chunks
        """
        if self.vectorstore is None:
            logger.error("Vector store not initialized!")
            return []

        # This converts your query to a vector and finds similar vectors
        results = self.vectorstore.similarity_search(query, k=k)

        return results

    def get_stats(self) -> Dict:
        """Get statistics about the vector store"""
        if self.vectorstore is None:
            return {"total_chunks": 0}

        count = self.vectorstore._collection.count()
        return {"total_chunks": count, "vectorstore_path": str(VECTORSTORE_DIR)}


# Test function
if __name__ == "__main__":
    """
    This runs when you execute this file directly.
    Useful for testing!
    """
    print("Testing Document Processor...")
    processor = DocumentProcessor()

    # Process all documents
    stats = processor.process_all_documents()

    # Test search
    if stats["total_chunks"] > 0:
        print("\n" + "=" * 50)
        print("Testing search...")
        results = processor.search("test query", k=2)
        print(f"Found {len(results)} results")
